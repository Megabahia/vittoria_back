import json

from .models import (
    ProductoImagen,
    Productos, ReporteAbastecimiento, ReporteStock, ReporteCaducidad, ReporteRotacion, IngresoProductos
)

from ...ADM.vittoria_integraciones.models import Integraciones
from ...ADM.vittoria_integraciones.serializers import IntegracionesSerializer

from .serializers import (
    DetallesSerializer, ProductosActualizarSerializer,
    ProductoCreateSerializer,
    ProductosSerializer, ProductosListSerializer,
    AbastecimientoListSerializer,
    StockListSerializer, CaducidadListSerializer, RotacionListSerializer, RefilListSerializer,
    HistorialAvisosSerializer, ImagenSerializer, PrediccionCrosselingSerializer,
    PrediccionRefilSerializer, PrediccionRefilOneSerializer, ProductoSearchSerializer,
    ListarProductoCreateSerializer, ProductosIntegracionesListSerializer
)
from rest_framework import status
from urllib.parse import urlparse
from ...ADM.vittoria_catalogo.models import Catalogo

from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from datetime import datetime
# excel
import openpyxl
from openpyxl import Workbook
from django.http import HttpResponse
# logs
from ...ADM.vittoria_logs.methods import createLog, datosTipoLog, datosProductosMDP
from .constantes import mapeoCrearProducto,mapeoActualizarProducto
from ...GDP.gdp_productos.serializers import ProductosSerializer

# declaracion variables log
datosAux = datosProductosMDP()
datosTipoLogAux = datosTipoLog()
# asignacion datos modulo
logModulo = datosAux['modulo']
logApi = datosAux['api']
# asignacion tipo de datos
logTransaccion = datosTipoLogAux['transaccion']
logExcepcion = datosTipoLogAux['excepcion']


# CRUD PRODUCTOS
# LISTAR TODOS
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def productos_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}

            if 'nombre' in request.data and request.data['nombre'] != '':
                filters['nombre__icontains'] = request.data['nombre']

            if 'codigoBarras' in request.data and request.data['codigoBarras'] != '':
                filters['codigoBarras__icontains'] = request.data['codigoBarras']

            if 'canalProducto' in request.data and request.data['canalProducto'] != '':
                filters['canal'] = request.data['canalProducto']

            if 'proveedor' in request.data and request.data['proveedor'] != '':
                filters['proveedor'] = request.data['proveedor']

            #Se realiza la exraccion de los canales de los productos por el motivo de
            #que existen productos con el mismo codigo pero con diferente canal y para realziar el filtro
            #al obtener un producto
            queryCanal = list(Productos.objects.values_list('canal', flat=True).distinct())

            # Serializar los datos

            query = Productos.objects.filter(**filters).order_by('-created_at')
            serializer = ProductosListSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(),
                                   'info': serializer.data,
                                   'canal': queryCanal}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)

        # ENCONTRAR UNO


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def productos_findOne(request, pk):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = Productos.objects.get(pk=pk, state=1)
        except Productos.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'GET':
            serializer = ProductoCreateSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# ENCONTRAR IMAGENES
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def producto_images_findOne(request, pk):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = ProductoImagen.objects.filter(producto=pk, state=1)
        except ProductoImagen.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'GET':
            serializer = DetallesSerializer(query, many=True)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# BORRAR IMAGENES
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def producto_images_delete(request, pk):
    nowDate = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'delete/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'BORRAR',
        'fechaInicio': str(nowDate),
        'dataEnviada': '{}',
        'fechaFin': str(nowDate),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = ProductoImagen.objects.get(pk=pk, state=1).delete()
        except ProductoImagen.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'DELETE':
            data = {'message': 'Se elimino correctamente.'}
            if data:
                createLog(logModel, data, logTransaccion)
                return Response(data, status=status.HTTP_200_OK)
            errors = {'message': 'No se elimino la imagen.'}
            createLog(logModel, errors, logExcepcion)
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# CREAR
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def productos_create(request):
    request.POST._mutable = True
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'create/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':

        logModel['dataEnviada'] = str(request.data)
        query = Productos.objects.filter(codigoBarras=request.data['codigoBarras'], state=1).first()
        if query is not None:
            errorNoExiste = {'error': 'Ya existe el producto'}
            createLog(logModel, errorNoExiste, logExcepcion)
            return Response(errorNoExiste, status=status.HTTP_404_NOT_FOUND)
        try:
            logModel['dataEnviada'] = str(request.data)
            request.data['created_at'] = str(timezone_now)
            if 'updated_at' in request.data:
                request.data.pop('updated_at')
            if 'parametrizacion' in request.data and request.data['parametrizacion'] == '0':
                request.data.pop('parametrizacion')
            if 'stockVirtual' in request.data and type(request.data['stockVirtual']) == 'str':
                stockVirtualTemporal = request.data.pop('stockVirtual')
                request.data['stockVirtual'] = json.dumps(stockVirtualTemporal)
            serializer = ProductoCreateSerializer(data=request.data)

            if serializer.is_valid():
                serializer.save()

                if (request.data['idPadre'] != ''):

                    productoPadre = Productos.objects.filter(codigoBarras=request.data['idPadre']).first()
                    productoHijo = Productos.objects.filter(codigoBarras=request.data['codigoBarras']).first()

                    if productoPadre and productoHijo:
                        productoPadre.stock = productoPadre.stock + int(request.data['stock'])
                        productoPadre.save()

                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)

# ACTUALIZAR
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def productos_update(request, pk):
    request.POST._mutable = True
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'update/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'ESCRIBIR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        logModel['dataEnviada'] = str(request.data)
        query = Productos.objects.filter(codigoBarras=request.data['codigoBarras'], canal=request.data['canal'], state=1).exclude(pk=pk).first()
        if query is not None:
            errorNoExiste = {'error': 'Ya existe el producto'}
            createLog(logModel, errorNoExiste, logExcepcion)
            return Response(errorNoExiste, status=status.HTTP_404_NOT_FOUND)
        try:
            logModel['dataEnviada'] = str(request.data)
            query = Productos.objects.get(pk=pk, state=1)
        except Productos.DoesNotExist:
            errorNoExiste = {'error': 'No existe'}
            createLog(logModel, errorNoExiste, logExcepcion)
            return Response(status=status.HTTP_404_NOT_FOUND)
        if request.method == 'POST':
            now = timezone.localtime(timezone.now())
            request.data['updated_at'] = str(now)
            if 'created_at' in request.data:
                request.data.pop('created_at')
            if 'parametrizacion' in request.data and request.data['parametrizacion'] == '0':
                request.data.pop('parametrizacion')
            if 'stockVirtual' in request.data and type(request.data['stockVirtual']) == 'str':
                stockVirtualTemporal = request.data.pop('stockVirtual')
                request.data['stockVirtual'] = json.dumps(stockVirtualTemporal)
            serializer = ProductosActualizarSerializer(query, data=request.data, partial=True)
            if ("idPadre" in request.data and request.data['idPadre'] != ''):

                productoPadre = Productos.objects.filter(codigoBarras=request.data['idPadre']).first()
                productoHijo = Productos.objects.filter(codigoBarras=request.data['codigoBarras']).first()


                if request.data['idPadre'] == '':
                    request.data['idPadre'] = ''

                if productoPadre and productoHijo:
                    if int(request.data['stock'])<productoHijo.stock:
                        diferenciaStock = productoHijo.stock-int(request.data['stock'])
                        productoHijo.stock = int(request.data['stock'])
                        productoHijo.save()
                        productoPadre.stock = productoPadre.stock - diferenciaStock
                        productoPadre.save()
                    else:
                        diferenciaStock = int(request.data['stock']) - productoHijo.stock
                        productoHijo.stock = int(request.data['stock'])
                        productoHijo.save()
                        productoPadre.stock = productoPadre.stock + diferenciaStock
                        productoPadre.save()

            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)

    # ELIMINAR


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def productos_delete(request, pk):
    nowDate = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'delete/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'BORRAR',
        'fechaInicio': str(nowDate),
        'dataEnviada': '{}',
        'fechaFin': str(nowDate),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = Productos.objects.get(pk=pk, state=1)
        except Productos.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
            return Response(status=status.HTTP_404_NOT_FOUND)
        productosTodos = Productos.objects.filter(idPadre=query.codigoBarras,state=1).exclude(pk=pk).all()

        if productosTodos:
            return Response('Este producto no puede eliminarse porque tiene productos hijos.', status=status.HTTP_404_NOT_FOUND)
        else:
            # tomar el dato
            if request.method == 'DELETE':
                serializer = ProductosSerializer(query, data={'state': '0', 'updated_at': str(nowDate)}, partial=True)
                if (query.idPadre != ''):
                    productoPadre = Productos.objects.filter(codigoBarras=query.idPadre).first()
                    if productoPadre and query:
                        productoPadre.stock = productoPadre.stock - query.stock
                        productoPadre.save()

                if serializer.is_valid():
                    serializer.save()
                    createLog(logModel, serializer.data, logTransaccion)
                    return Response(serializer.data, status=status.HTTP_200_OK)
                createLog(logModel, serializer.errors, logExcepcion)
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)

    # SEARCH PRODUCTO


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def search_producto_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}
            if 'codigoBarras' in request.data:
                if request.data['codigoBarras'] != '':
                    filters['codigoBarras__icontains'] = str(request.data['codigoBarras'])
            if 'nombre' in request.data:
                if request.data['nombre'] != '':
                    filters['nombre__icontains'] = str(request.data['nombre'])

            # Serializar los datos
            query = Productos.objects.filter(**filters).order_by('-created_at')
            serializer = ProductosListSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(),
                                   'info': serializer.data}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)

        # SEARCH PRODUCTO


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def search_producto_codigo_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'search/producto/codigo/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            if request.data['canalProducto'] is None or request.data['canalProducto'] == '':
                createLog(logModel, 'Seleccione un canal', logExcepcion)
                return Response('Seleccione un canal', status=status.HTTP_404_NOT_FOUND)

            filters = {
                'codigoBarras': request.data['codigoBarras'],
                'state': 1,
                'estado': 'Activo',
                'stockVirtual__contains': {'canal': request.data['canalProducto'], 'estado': True}
            }

            query = Productos.objects.filter(**filters).first()
            # Verifica si el objeto existe antes de aplicar más filtros
            if query:
                url_completa = request.data['canal']
                valorUnitario = float(request.data['valorUnitario'])

                indice_com = url_completa.find('.com')

                url_cortada = url_completa[:indice_com + 4] if indice_com != -1 else url_completa

                if url_cortada=='Landing-Producto' or url_cortada=='Landing-Page':
                    query.precio = query.precioLandingOferta if query.precioLandingOferta is not None else 0
                    query.mensaje = ""
                elif url_cortada=='todomegacentro.megadescuento.com':
                    if valorUnitario == 0:
                        if(query.precioVentaA!=0):
                            query.precio = query.precioVentaA if query.precioVentaA is not None else 0
                            query.mensaje = ""
                        else:
                            query.precio = query.precioVentaB if query.precioVentaB is not None else 0
                            query.mensaje = ""
                    elif query.precioVentaA == valorUnitario:
                        query.precio=query.precioVentaA
                        query.mensaje = ""
                    elif query.precioVentaB == valorUnitario:
                        query.precio=query.precioVentaB
                        query.mensaje = ""
                    else:
                        query.precio=valorUnitario
                        query.mensaje="NO COINCIDE CON OFERTA TODOMEGACENTRO"
                elif url_cortada=='mayorista.megadescuento.com':
                    if valorUnitario == 0:
                        if (query.precioVentaC != 0):
                            query.precio = query.precioVentaC if query.precioVentaC is not None else 0
                            query.mensaje = ""
                        elif (query.precioVentaD != 0):
                            query.precio = query.precioVentaD if query.precioVentaD is not None else 0
                            query.mensaje = ""
                        else:
                            query.precio = query.precioVentaE if query.precioVentaE is not None else 0
                            query.mensaje = ""

                        valorUnitario = query.precio

                    elif query.precioVentaC == valorUnitario:
                        query.precio = query.precioVentaC
                        query.mensaje = ""
                    elif query.precioVentaD == valorUnitario:
                        query.precio = query.precioVentaD
                        query.mensaje = ""

                    elif query.precioVentaE == valorUnitario:
                        query.precio = query.precioVentaE
                        query.mensaje = ""
                    else:
                        query.precio = valorUnitario
                        query.mensaje ="NO COINCIDE CON OFERTA MAYORISTA"
                elif url_cortada=='contraentrega.megadescuento.com':
                    if valorUnitario == 0:
                        query.precio = query.precioVentaBultos if query.precioVentaBultos is not None else 0
                        query.mensaje = ""
                    elif query.precioVentaBultos == valorUnitario:
                        query.precio = query.precioVentaBultos
                        query.mensaje = ""
                    else:
                        query.precio = valorUnitario
                        query.mensaje ="NO COINCIDE CON LA OFERTA CONTRAENTREGA"
                elif url_cortada=='megadescuento.com':
                    if valorUnitario == 0:
                        query.precio = query.precioOferta if query.precioOferta is not None else 0
                        query.mensaje = ""
                    elif query.precioOferta == valorUnitario:
                        query.precio = query.precioOferta
                        query.mensaje = ""
                    else:
                        query.precio = valorUnitario
                        query.mensaje ="NO COINCIDE CON OFERTA MEGADESCUENTO"
                else:
                    if valorUnitario == 0:
                        query.precio = query.precioVentaF if query.precioVentaF is not None else 0
                        query.mensaje = ""
                    elif query.precioVentaF == valorUnitario:
                        query.precio = query.precioVentaF
                        query.mensaje = ""
                    else:
                        query.precio=valorUnitario
                        query.mensaje ="NO EXISTE OFERTA SIMILAR AL PRECIO"
                        query.mensaje =''

                query2 = Productos.objects.filter(id=query.id, envioNivelNacional=False).first()
                # Este if sirve para validar que el producto que se configuro para una ciudad en especifico y no puede ser vendido a otra ciudad
                if query2:
                    if 'lugarVentaCiudad' in request.data and request.data['lugarVentaCiudad'] != '':
                        filters['lugarVentaCiudad'] = request.data['lugarVentaCiudad']
                        query = Productos.objects.filter(**filters).first()
                        query.precio = query.precioLandingOferta
                        query.mensaje = ""
                    else:
                        query = query2
                        query.precio = query2.precioLandingOferta
                        query.mensaje = ""
        except Productos.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'POST':
            serializer = ProductoSearchSerializer(query)

            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def search_producto_codigo_canal_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'search/producto/codigo/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            '''filters = {
                'codigoBarras': request.data['codigoBarras'],
                'state': 1,
                'estado': 'Activo',
                #'stockVirtual__contains': {'canal': request.data['canalProducto'], 'estado': True}
            }'''

            filters = {}

            if 'canal' in request.data and request.data['canal'] != '':
                filters['canal'] = request.data['canal']

            if 'state' in request.data and request.data['state'] != '':
                filters['state'] = request.data['state']

            if 'estado' in request.data and request.data['estado'] != '':
                filters['estado'] = request.data['estado']

            if 'codigoBarras' in request.data and request.data['codigoBarras'] != '':
                filters['codigoBarras'] = request.data['codigoBarras']

            if 'nombre' in request.data and request.data['nombre'] != '':
                filters['nombre__icontains'] = request.data['nombre']

            query = Productos.objects.filter(**filters).first()
            query2 = Productos.objects.filter(**filters)

            if(query is None or query2 is None):
                return Response('El producto no existe', status=status.HTTP_404_NOT_FOUND)
            # Verifica si el objeto existe antes de aplicar más filtros

        except Productos.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'POST':

            serializer = ProductosListSerializer(query)
            serializer2 = ProductosIntegracionesListSerializer(query2, many=True)

            queryParamsCanal = Integraciones.objects.filter(valor = serializer.data['canal']).first()
            serializer_canal = IntegracionesSerializer(queryParamsCanal)

            new_serializer_data = {'producto': serializer.data,
                                   'productos': serializer2.data,
                                   'integraciones_canal': serializer_canal.data}

            createLog(logModel, new_serializer_data, logTransaccion)
            return Response(new_serializer_data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)

# ABASTECIMIENTO
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def abastecimiento_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'abastecimiento/list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}
            if request.data['inicio'] != '':
                filters['fechaMaximaStock__gte'] = str(request.data['inicio'])
            if request.data['fin'] != '':
                filters['fechaMaximaStock__lte'] = str(request.data['fin'])
            if 'categoria' in request.data:
                if request.data['categoria'] != '':
                    filters['producto__categoria__icontains'] = str(request.data['categoria'])
            if 'subCategoria' in request.data:
                if request.data['subCategoria'] != '':
                    filters['producto__subCategoria__icontains'] = str(request.data['subCategoria'])

                    # Serializar los datos
            query = ReporteAbastecimiento.objects.filter(**filters).order_by('-created_at')
            serializer = AbastecimientoListSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(),
                                   'info': serializer.data}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


# STOCK
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def stock_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'caducidad/list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}
            if request.data['inicio'] != '':
                filters['fechaUltimaStock__gte'] = str(request.data['inicio'])
            if request.data['fin'] != '':
                filters['fechaUltimaStock__lte'] = str(request.data['fin'])
            if 'categoria' in request.data:
                if request.data['categoria'] != '':
                    filters['producto__categoria__icontains'] = str(request.data['categoria'])
            if 'subCategoria' in request.data:
                if request.data['subCategoria'] != '':
                    filters['producto__subCategoria__icontains'] = str(request.data['subCategoria'])

                    # Serializar los datos
            query = ReporteStock.objects.filter(**filters).order_by('-created_at')
            serializer = StockListSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(),
                                   'info': serializer.data}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


# CADUCIDAD
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def caducidad_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'caducidad/list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}
            if request.data['inicio'] != '':
                filters['fechaCaducidad__gte'] = str(request.data['inicio'])
            if request.data['fin'] != '':
                filters['fechaCaducidad__lte'] = str(request.data['fin'])
            if 'categoria' in request.data:
                if request.data['categoria'] != '':
                    filters['producto__categoria__icontains'] = str(request.data['categoria'])
            if 'subCategoria' in request.data:
                if request.data['subCategoria'] != '':
                    filters['producto__subCategoria__icontains'] = str(request.data['subCategoria'])

                    # Serializar los datos
            query = ReporteCaducidad.objects.filter(**filters).order_by('-created_at')
            serializer = CaducidadListSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(),
                                   'info': serializer.data}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


# ROTACION
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def rotacion_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'rotacion/list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}
            if request.data['inicio'] != '':
                filters['fechaInicio__gte'] = str(request.data['inicio'])
            if request.data['fin'] != '':
                filters['fechaFin__lte'] = str(request.data['fin'])
            if 'categoria' in request.data:
                if request.data['categoria'] != '':
                    filters['producto__categoria__icontains'] = str(request.data['categoria'])
            if 'subCategoria' in request.data:
                if request.data['subCategoria'] != '':
                    filters['producto__subCategoria__icontains'] = str(request.data['subCategoria'])

                    # Serializar los datos
            query = ReporteRotacion.objects.filter(**filters).order_by('-created_at')
            serializer = RotacionListSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(),
                                   'info': serializer.data}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)

        # REFIL


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def refil_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'refil/list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}
            if request.data['inicio'] != '' and request.data['fin'] != '':
                filters['refil__range'] = [int(request.data['inicio']), int(request.data['fin'])]
            elif request.data['inicio'] != '':
                filters['refil__gte'] = str(request.data['inicio'])
            elif request.data['fin'] != '':
                filters['refil__lte'] = str(request.data['fin'])
            if 'categoria' in request.data:
                if request.data['categoria'] != '':
                    filters['categoria__icontains'] = str(request.data['categoria'])
            if 'subCategoria' in request.data:
                if request.data['subCategoria'] != '':
                    filters['subCategoria__icontains'] = str(request.data['subCategoria'])

                    # Serializar los datos
            query = Productos.objects.filter(**filters).order_by('-created_at')
            serializer = RefilListSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(),
                                   'info': serializer.data}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)

        # METODO SUBIR ARCHIVOS EXCEL


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uploadEXCEL_crearProductos(request):
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'POST':
            first = True  # si tiene encabezado
            uploaded_file = request.FILES['documento']
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(uploaded_file)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Hoja1"]
            lines = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            lines.append(row_data)

        for dato in lines:
            contTotal += 1
            if first:
                first = False
                continue
            else:
                if worksheet.iter_cols():
                    resultadoInsertar = insertarDato_Producto(dato)
                    if resultadoInsertar != 'Dato insertado correctamente':
                        if resultadoInsertar in 'Codigo producto':
                            contInvalidos += 1
                            errores.append(
                                {"error": "Producto no encontrado " + str(contTotal) + ": " + str(resultadoInsertar)})
                        else:
                            contInvalidos += 1
                            errores.append(
                                {"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                    else:
                        contValidos += 1
                else:
                    contInvalidos += 1
                    errores.append({"error": "Error en la línea " + str(
                        contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

        result = {"mensaje": "La Importación se Realizo Correctamente",
                  "correctos": contValidos,
                  "incorrectos": contInvalidos,
                  "errores": errores
                  }
        return Response(result, status=status.HTTP_201_CREATED)

    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# INSERTAR DATOS EN LA BASE INDIVIDUAL
def insertarDato_Producto(dato):
    try:
        timezone_now = timezone.localtime(timezone.now())
        data = {}
        data['codigoBarras'] = dato[0].replace('"', "") if dato[0].replace('"', "") != "NULL" else None
        data['descripcion'] = dato[1].replace('"', "") if dato[1] != "NULL" else None
        data['stock'] = dato[2].replace('"', "") if dato[2] != "NULL" else None
        data['lote'] = dato[3].replace('"', "") if dato[3] != "NULL" else None
        data['fechaElaboracion'] = str(dato[4].replace('"', "")[:10]) if dato[4] != "NULL" else None
        data['fechaCaducidad'] = str(dato[5].replace('"', "")[:10]) if dato[5] != "NULL" else None
        data['costoCompra'] = str(dato[6].replace('"', "")) if dato[6] != "NULL" else None
        data['updated_at'] = str(timezone_now)
        # inserto el dato con los campos requeridos
        query = Productos.objects.get(codigoBarras=data['codigoBarras'])
        for key, value in data.items():
            setattr(query, key, value)
        query.save()
        if query == 0:
            return 'Codigo producto %(code)s no existe' % {"code": data['codigoBarras']}
        # CREAR INGRESO PRODUCTOS
        IngresoProductos.objects.create(cantidad=data['stock'], fechaElaboracion=str(data['fechaElaboracion']),
                                        fechaCaducidad=str(data['fechaCaducidad']), precioCompra=data['costoCompra'],
                                        producto=query)
        # CREAR REPORTE STOCK
        ReporteStock.objects.create(fechaUltimaStock=datetime.today().strftime('%Y-%m-%d'),
                                    montoCompra=data['costoCompra'], producto=query)
        return 'Dato insertado correctamente'
    except Exception as e:
        return str(e)


# CREAR ABASTECIMIENTO
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def abastecimiento_create(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'abastecimiento/create/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # request.data['created_at'] = str(timezone_now)
            # if 'updated_at' in request.data:
            #     request.data.pop('updated_at')

            serializer = HistorialAvisosSerializer(data=request.data['data'], many=True)
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


# OBTENER URL IMAGEN
@api_view(['POST'])
def productoImagen_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'producto/image/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = ProductoImagen.objects.filter(producto__codigoBarras=str(request.data['codigo']),
                                                  producto__state=1).first()
        except ProductoImagen.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'POST':
            serializer = ImagenSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# OBTENER PREDICCION CROSSELING
@api_view(['POST'])
def prediccion_crosseling_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'prediccionCrosseling/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = Productos.objects.filter(codigoBarras=str(request.data['codigo']), state=1).first()
        except Productos.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'POST':
            query = ReporteRotacion.objects.filter(producto__subCategoria=query.subCategoria, tipoRotacion='Bajo',
                                                   state=1).order_by('-created_at', '-producto__stock')
            serializer = PrediccionCrosselingSerializer(query[0:3], many=True)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# OBTENER PRODUCTO REFIL
@api_view(['POST'])
def producto_refil_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'prediccionCrosseling/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = Productos.objects.filter(codigoBarras=request.data['producto'], state=1)
        except Productos.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'POST':
            serializer = PrediccionRefilSerializer(query, many=True)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# OBTENER PREDICCION REFIL
@api_view(['POST'])
def prediccion_refil_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'prediccionRefil/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = ProductoImagen.objects.filter(producto__codigoBarras=request.data['producto'], state=1).first()
        except ProductoImagen.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'POST':
            serializer = PrediccionRefilOneSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# OBTENER PREDICCION PRODUCTOS NUEVOS
@api_view(['POST'])
def prediccion_productosNuevos_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'prediccionProductosNuevos/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = ProductoImagen.objects.filter(producto__codigoBarras=request.data['producto'], state=1).first()
        except ProductoImagen.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'POST':
            query = ReporteRotacion.objects.filter(producto__categoria=query.producto.categoria, tipoRotacion='Bajo',
                                                   state=1).order_by('-created_at', '-producto__stock')
            serializer = PrediccionCrosselingSerializer(query[0:3], many=True)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
def productos_findOne_free(request, pk):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            filters = {
                'pk': pk,
                'state': 1
            }
            if 'estado' in request.data and request.data['estado'] != '':
                filters['estado'] = request.data['estado']
            if 'estadoLanding' in request.data and request.data['estadoLanding'] != '':
                filters['estadoLanding'] = request.data['estadoLanding']
            query = Productos.objects.get(**filters)
        except Productos.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'POST':
            serializer = ListarProductoCreateSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
def productos_findOne_codigo_producto(request, pk):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = Productos.objects.get(codigoBarras=pk, state=1)
        except Productos.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'GET':
            serializer = ProductoCreateSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def productos_exportar(request):
    """
    Este metodo realiza una exportacion de todos los productos en excel de la tabla productos, de la base de datos mdp
    @rtype: DEvuelve un archivo excel
    """
    # Obtener los datos que deseas incluir en el archivo Excel
    filters = {"state": "1"}

    nombre = request.GET.get('nombre', None)
    codigoBarras = request.GET.get('codigoBarras', None)
    if nombre is not None:
        filters['nombre__icontains'] = nombre

    if codigoBarras is not None:
        filters['codigoBarras__icontains'] = codigoBarras

    # Serializar los datos
    query = Productos.objects.filter(**filters).order_by('-created_at')

    # Crear un libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active

    # Definir los encabezados
    ws.append(['CÓDIGO DE BARRAS', 'NOMBRE', 'CATEGORÍA', 'SUBCATEGORÍA', 'STOCK', 'ESTADO'])

    # Agregar datos de productos a las filas siguientes
    for producto in query:
        productoSerializer = ProductosListSerializer(producto).data

        # Verificar si los campos están vacíos antes de acceder a ellos
        codigo_barras = productoSerializer.get('codigoBarras', '')
        nombre = productoSerializer.get('nombre', '')
        categoria = productoSerializer.get('categoria', '')
        subcategoria = productoSerializer.get('subCategoria', '')
        stock = productoSerializer.get('stock', '')
        estado = productoSerializer.get('estado', '')

        # Agregar datos a la hoja de trabajo
        ws.append([codigo_barras, nombre, categoria, subcategoria, stock, estado])

    # Crear una respuesta HTTP con el contenido del libro de trabajo
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_productos.xlsx"'
    wb.save(response)
    return response


@api_view(['POST'])
def productos_create_woocommerce(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'create/woocommerce',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)

            #dominio_completo = request.headers.get('X-Wc-Webhook-Source')
            # Utiliza urlparse para obtener la información de la URL
            #parsed_url = urlparse(dominio_completo)
            # Combina el nombre de host (dominio) y el esquema (protocolo)
            #domain = parsed_url.netloc
            #dominio_permitidos = Catalogo.objects.filter(tipo='INTEGRACION_WOOCOMMERCE', valor=domain).first()
            #if dominio_permitidos is None:
            #    error = f"Llego un dominio: {domain}"
            #    createLog(logModel, error, logTransaccion)
            #    return Response(error, status=status.HTTP_400_BAD_REQUEST)

            data=mapeoCrearProducto(request)
            serializer = ProductoCreateSerializer(data = data)

            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            err = {"error": "Un error ha ocurrido: {}".format(e)}
            createLog(logModel,err,logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
def productos_update_woocommerce(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'update/woocommerce',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'ESCRIBIR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:

        logModel['dataEnviada'] = str(request.data)

        #dominio_completo = request.headers.get('X-Wc-Webhook-Source')
        # Utiliza urlparse para obtener la información de la URL
        #parsed_url = urlparse(dominio_completo)
        # Combina el nombre de host (dominio) y el esquema (protocolo)
        #domain = parsed_url.netloc
        #dominio_permitidos = Catalogo.objects.filter(tipo='INTEGRACION_WOOCOMMERCE', valor=domain).first()
        #if dominio_permitidos is None:
        #    error = f"Llego un dominio: {domain}"
        #    createLog(logModel, error, logTransaccion)
        #    return Response(error, status=status.HTTP_400_BAD_REQUEST)

        index = request.data['permalink'].find('.com')
        if index != -1:
            canal = request.data['permalink'][:index + 4]
        else:
            canal = request.data['permalink']

        canal = canal.replace('https://', '')

        query = Productos.objects.filter(codigoBarras=request.data['sku'], canal=canal, state=1).exclude(codigoBarras=request.data['sku'],
                                                                       canal=canal).first()
        print(query)
        if query is not None:
            errorNoExiste = {'error': 'Ya existe el producto'}
            createLog(logModel, errorNoExiste, logExcepcion)
            return Response(errorNoExiste, status=status.HTTP_404_NOT_FOUND)
        try:
            logModel['dataEnviada'] = str(request.data)
            query = Productos.objects.get(codigoBarras=request.data['sku'], canal=canal, state=1)
        except Productos.DoesNotExist:
            errorNoExiste = {'error': 'No existe'}
            createLog(logModel, errorNoExiste, logExcepcion)
            return Response(status=status.HTTP_404_NOT_FOUND)
        if request.method == 'POST':
            now = timezone.localtime(timezone.now())
            request.data['updated_at'] = str(now)
            if 'created_at' in request.data:
                request.data.pop('created_at')

            data = mapeoActualizarProducto(request)

            serializer = ProductosActualizarSerializer(query, data=data, partial=True)

            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
def productos_delete_woocommerce(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'delete/woocommerce',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'BORRAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:

        dominio_completo = request.headers.get('X-Wc-Webhook-Source')
        # Utiliza urlparse para obtener la información de la URL
        parsed_url = urlparse(dominio_completo)
        # Combina el nombre de host (dominio) y el esquema (protocolo)
        domain = parsed_url.netloc
        dominio_permitidos = Catalogo.objects.filter(tipo='INTEGRACION_WOOCOMMERCE', valor=domain).first()
        if dominio_permitidos is None:
            error = f"Llego un dominio: {domain}"
            createLog(logModel, error, logTransaccion)
            return Response(error, status=status.HTTP_400_BAD_REQUEST)

        try:
            query = Productos.objects.get(woocommerceId=request.data['id'], state=1)
        except Productos.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)

        if request.method == 'DELETE':

            query.estado ='Inactivo'
            query.state =0
            query.save()
            createLog(logModel, 'Producto retirado de la lista', logExcepcion)
            return Response('Producto retirado de la lista', status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
def productos_restore_woocommerce(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'restore/woocommerce',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'ESCRIBIR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        logModel['dataEnviada'] = str(request.data)

        dominio_completo = request.headers.get('X-Wc-Webhook-Source')
        # Utiliza urlparse para obtener la información de la URL
        parsed_url = urlparse(dominio_completo)
        # Combina el nombre de host (dominio) y el esquema (protocolo)
        domain = parsed_url.netloc
        dominio_permitidos = Catalogo.objects.filter(tipo='INTEGRACION_WOOCOMMERCE', valor=domain).first()
        if dominio_permitidos is None:
            error = f"Llego un dominio: {domain}"
            createLog(logModel, error, logTransaccion)
            return Response(error, status=status.HTTP_400_BAD_REQUEST)

        index = request.data['permalink'].find('.com')
        if index != -1:
            canal = request.data['permalink'][:index + 4]
        else:
            canal = request.data['permalink']

        canal = canal.replace('https://', '')

        try:
            query = Productos.objects.filter(codigoBarras=request.data['sku'], woocommerceId=request.data['id'],
                                             canal=canal).first()

            if request.method == 'POST':
                query.estado = 'Activo'
                query.state = 1

                query.save()
                createLog(logModel, 'Se ha restaurado el producto', logExcepcion)
                return Response('Producto restaurado', status=status.HTTP_200_OK)
        except Productos.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)

    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)