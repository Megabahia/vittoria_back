from django.http import HttpResponse
from ...config.config import aws_s3_instancia

from .models import (
    Productos, ProductosImagenes
)
from ..mdp_productos.models import (
    Productos as ProductosMDP, ProductoImagen as ProductosImagenesMDP
)
from .serializers import (
    ProductosSerializer, ArchivosFacturasSerializer, ProveedoresSerializer, ProductosResource, ProductosStockResource
)
from .utils import enviarCorreoInventario
from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from datetime import datetime
# excel
import openpyxl
import pandas as pd
import openpyxl
# logs
from ...ADM.vittoria_logs.methods import createLog, datosTipoLog, datosProductosMDP

# declaracion variables log
datosAux = datosProductosMDP()
datosTipoLogAux = datosTipoLog()
# asignacion datos modulo
logModulo = datosAux['modulo']
logApi = datosAux['api']
# asignacion tipo de datos
logTransaccion = datosTipoLogAux['transaccion']
logExcepcion = datosTipoLogAux['excepcion']


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uploadEXCEL_subirProductosProveedores(request):
    """
    El metodo sirve para crear
    @type request: REcibe los campos de la tabla credito archivo
    @rtype: DEvuelve el registro creado
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'create/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            request.data['created_at'] = str(timezone_now)
            if 'updated_at' in request.data:
                request.data.pop('updated_at')

            serializer = ArchivosFacturasSerializer(data=request.data)
            result = uploadEXCEL_crearProductos(request)
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                return Response(result, status=status.HTTP_201_CREATED)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


def uploadEXCEL_crearProductos(request):
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'POST':
            first = True  # si tiene encabezado
            uploaded_file = request.FILES['archivo']
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(uploaded_file)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Hoja1"]
            lines = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            lines.append(row_data)

        for dato in lines:
            contTotal += 1
            if first:
                first = False
                continue
            else:
                if worksheet.iter_cols():
                    resultadoInsertar = insertarDato_Factura(dato)
                    if resultadoInsertar != 'Dato insertado correctamente':
                        if resultadoInsertar in 'Codigo producto':
                            contInvalidos += 1
                            errores.append(
                                {"error": "Producto no encontrado " + str(contTotal) + ": " + str(resultadoInsertar)})
                        else:
                            contInvalidos += 1
                            errores.append(
                                {"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                    else:
                        contValidos += 1
                else:
                    contInvalidos += 1
                    errores.append({"error": "Error en la línea " + str(
                        contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

        result = {"mensaje": "La Importación se Realizo Correctamente",
                  "correctos": contValidos,
                  "incorrectos": contInvalidos,
                  "errores": errores
                  }
        return result

    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# INSERTAR DATOS EN LA BASE INDIVIDUAL
def insertarDato_Factura(dato):
    try:
        timezone_now = timezone.localtime(timezone.now())
        if dato[1] == None and dato[6] == None or dato[1] == 'None' and dato[6] == 'None':
            return 'Dato insertado correctamente'
        facturaEncabezadoQuery = ProductosMDP.objects.filter(codigoBarras=dato[1]).first()
        if facturaEncabezadoQuery is None:
            facturaEncabezado = {}
            facturaEncabezado['codigoBarras'] = dato[1].replace('"', "") if dato[1] != "NULL" else None
            # Llamar a la función obtener_archivo con una clave específica
            fotoFrente = aws_s3_instancia.get_foto_frente_url(facturaEncabezado['codigoBarras'])
            fotoBonita = aws_s3_instancia.get_foto_bonita_url(facturaEncabezado['codigoBarras'])
            fotoOriginal = aws_s3_instancia.get_foto_original_url(facturaEncabezado['codigoBarras'])
            facturaEncabezado['nombre'] = dato[3].replace('"', "") if dato[3] != "NULL" else None
            facturaEncabezado['descripcion'] = dato[5].replace('"', "") if dato[5] != "NULL" else None
            facturaEncabezado['precioVentaA'] = round(float(dato[7].replace('"', "") if dato[7] != "NULL" else 0), 2)
            facturaEncabezado['estado'] = 'Activo'
            facturaEncabezado['stock'] = 0
            facturaEncabezado['created_at'] = str(timezone_now)
            facturaEncabezado['updated_at'] = str(timezone_now)
            facturaEncabezado['precioOferta'] = round(float(dato[7].replace('"', "") if dato[7] != "NULL" else 0), 2)
            facturaEncabezado['fechaElaboracion'] = str(timezone_now)[:10]
            facturaEncabezado['fechaCaducidad'] = str(timezone_now)[:10]
            facturaEncabezado['proveedor'] = dato[6].replace('"', "") if dato[6] != "NULL" else None
            producto = ProductosMDP.objects.create(**facturaEncabezado)
            if fotoOriginal:
                ProductosImagenes.objects.create(**{
                    "producto": producto.id,
                    "imagen": fotoOriginal,
                    "created_at": str(timezone_now)
                })
            if fotoBonita:
                ProductosImagenes.objects.create(**{
                    "producto": producto.id,
                    "imagen": fotoBonita,
                    "created_at": str(timezone_now)
                })
            if fotoFrente:
                ProductosImagenes.objects.create(**{
                    "producto": producto.id,
                    "imagen": fotoFrente,
                    "created_at": str(timezone_now)
                })

        return 'Dato insertado correctamente'
    except Exception as e:
        print('error', e)
        return str(e)


# CRUD PRODUCTOS
# LISTAR TODOS
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def proveedores_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}

            if 'codigoBarras' in request.data and request.data['codigoBarras'] != '':
                filters['codigoBarras__icontains'] = request.data['codigoBarras']

            if 'proveedor' in request.data and request.data['proveedor'] != '':
                filters['proveedor'] = request.data['proveedor']

            # Serializar los datos
            query = Productos.objects.filter(**filters).order_by('-created_at')
            serializer = ProductosSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(),
                                   'info': serializer.data}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def proveedores_list_distinct(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'GET':
        try:
            logModel['dataEnviada'] = str(request.data)

            # Serializar los datos
            query = ProductosMDP.objects.values('proveedor').distinct()
            serializer = ProveedoresSerializer(query, many=True)
            new_serializer_data = serializer.data
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def productos_exportar(request):
    """
    Este metodo realiza una exportacion de todos los catalogos en excel de la tabla catalogo, de la base de datos central
    @rtype: DEvuelve un archivo excel
    """
    person_resource = ProductosResource()
    dataset = person_resource.export()
    response = HttpResponse(dataset.xls, content_type="application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename="foo.xls"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def generar_productos_stock_exportar(request):
    """
    Este metodo realiza una exportacion de todos los catalogos en excel de la tabla catalogo, de la base de datos central
    @rtype: DEvuelve un archivo excel
    """
    person_resource = ProductosStockResource()
    dataset = person_resource.export()
    response = HttpResponse(dataset.xls, content_type="application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename="foo.xls"'
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def productos_cargar_stock(request):
    """
    El metodo sirve para crear
    @type request: REcibe los campos de la tabla credito archivo
    @rtype: DEvuelve el registro creado
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'create/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            request.data['created_at'] = str(timezone_now)
            if 'updated_at' in request.data:
                request.data.pop('updated_at')

            serializer = ArchivosFacturasSerializer(data=request.data)
            result = uploadEXCEL_stockProductos(request)
            if serializer.is_valid():
                serializer.save()
                enviarCorreoInventario(request.user, result)

                createLog(logModel, serializer.data, logTransaccion)
                return Response(result, status=status.HTTP_201_CREATED)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def productos_cargar_stock_megabahia(request):
    """
    El metodo sirve para crear
    @type request: REcibe los campos de la tabla credito archivo
    @rtype: DEvuelve el registro creado
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'create/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            request.data['created_at'] = str(timezone_now)
            if 'updated_at' in request.data:
                request.data.pop('updated_at')
            serializer = ArchivosFacturasSerializer(data=request.data)
            result = uploadEXCEL_stockProductosMegabahia(request)
            if serializer.is_valid():
                serializer.save()
                enviarCorreoInventario(request.user, result)
                createLog(logModel, result, logTransaccion)
                return Response(result, status=status.HTTP_201_CREATED)
            createLog(logModel, result, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


def uploadEXCEL_stockProductos(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': 'mdp/gestion-inventario/cargar/stock/megadescuento',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'POST':
            first = True  # si tiene encabezado
            uploaded_file = request.FILES['archivo']
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(uploaded_file)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Hoja1"]
            lines = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            lines.append(row_data)

        for dato in lines:
            contTotal += 1
            if first:
                first = False
                continue
            else:
                if worksheet.iter_cols():
                    resetearStock = True if 'resetearStock' in request.data and request.data['resetearStock'] == 'true' else False
                    resultadoInsertar = insertarDato_StockProducto(dato, resetearStock)
                    if resultadoInsertar != 'Dato insertado correctamente':
                        if resultadoInsertar in 'Codigo producto':
                            contInvalidos += 1
                            errores.append(
                                {"error": "Producto no encontrado " + str(contTotal) + ": " + str(resultadoInsertar)})
                        else:
                            contInvalidos += 1
                            errores.append(
                                {"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                            createLog(logModel,
                                      {"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)},
                                      logTransaccion)
                    else:
                        createLog(logModel, resultadoInsertar, logTransaccion)
                        contValidos += 1
                else:
                    contInvalidos += 1
                    errores.append({"error": "Error en la línea " + str(
                        contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

        result = {"mensaje": "La Importación se Realizo Correctamente",
                  "correctos": contValidos,
                  "incorrectos": contInvalidos,
                  "errores": errores
                  }
        return result

    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)

def uploadEXCEL_stockProductosMegabahia(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': 'mdp/gestion-inventario/cargar/stock/megabahia',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'POST':
            first = True  # si tiene encabezado
            archivo_xls = request.FILES['archivo']
            archivo_xlsx = 'archivo_convertido.xlsx'
            # Utiliza pandas para leer el archivo .xls y guardarlo como .xlsx
            df = pd.read_excel(archivo_xls, dtype=str)
            df.to_excel(archivo_xlsx, index=False)
            # listaCodigos = df['Cód. Producto'].tolist()
            # # print('listaCodigos', listaCodigos)
            # cantidades = pd.to_numeric(df['Existencia Total'])
            # cantidadesList = cantidades.tolist()
            # # print(cantidadesList)
            # # Realizar la consulta SELECT IN y actualizar los valores de stock
            # # codes = ['IWO-relojMD', 'TCL30SEMD3', 'sudadera-LIFEMD']
            # # amount = [50, 50, 50]
            # productos_update = []
            # for codigo, cantidad in zip(listaCodigos, cantidadesList):
            #     producto = ProductosMDP.objects.filter(codigoBarras=codigo).first()
            #     if producto:
            #         producto.stock = cantidad
            #         productos_update.append(producto)
            #
            # ProductosMDP.objects.bulk_update(productos_update, fields=['stock'])
            # # for producto in productos:
            # #     print('se guardo', producto.codigoBarras)
            # #     cantidad = int(df[df["Cód. Producto"] == producto.codigoBarras].iloc[0, 3])
            # #     producto.stock = producto.stock + cantidad if cantidad >= 0 else 0
            # #     producto.save()
            # print('si guardo')
            # return {'prueba'}
            # Carga del archivo .xlsx con openpyxl
            wb = openpyxl.load_workbook(archivo_xlsx)
            # Seleccionar una hoja específica del libro, por ejemplo, la primera hoja
            worksheet = wb.active
            # Crear una lista para almacenar los datos
            lines = list()
        # Iterar sobre las filas y columnas de la hoja
        for row in worksheet.iter_rows(values_only=True):
            row_data = list()
            for cell in row:
                row_data.append(str(cell))
            lines.append(row_data)

        for dato in lines:
            contTotal += 1
            if first:
                first = False
                continue
            else:
                if worksheet.iter_cols():
                    resetearStock = True if 'resetearStock' in request.data and request.data['resetearStock'] == 'true' else False
                    logModel['dataEnviada'] = str(dato)
                    resultadoInsertar = insertarDato_StockProductoMegabahia(dato, resetearStock)
                    if resultadoInsertar != 'Dato insertado correctamente':
                        if resultadoInsertar in 'Codigo producto':
                            contInvalidos += 1
                            errores.append(
                                {"error": "Producto no encontrado " + str(contTotal) + ": " + str(resultadoInsertar)})
                        else:
                            contInvalidos += 1
                            errores.append(
                                {"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                        createLog(logModel, {"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)}, logTransaccion)
                    else:
                        createLog(logModel, resultadoInsertar, logTransaccion)
                        contValidos += 1
                else:
                    contInvalidos += 1
                    errores.append({"error": "Error en la línea " + str(
                        contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

        result = {"mensaje": "La Importación se Realizo Correctamente",
                  "correctos": contValidos,
                  "incorrectos": contInvalidos,
                  "errores": errores
                  }
        return result

    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)




# INSERTAR DATOS EN LA BASE INDIVIDUAL
def insertarDato_StockProducto(dato, resetearStock):
    try:
        timezone_now = timezone.localtime(timezone.now())
        if dato[1] == None and dato[6] == None or dato[1] == 'None' and dato[6] == 'None':
            return 'Dato insertado correctamente'
        facturaEncabezadoQuery = ProductosMDP.objects.filter(codigoBarras=dato[1]).first()
        if facturaEncabezadoQuery and None is not facturaEncabezadoQuery:
            if resetearStock:
                facturaEncabezadoQuery.stock = int(dato[4].replace('"', "") if dato[4] != "NULL" else 0)
            else:
                facturaEncabezadoQuery.stock = facturaEncabezadoQuery.stock + int(
                    dato[4].replace('"', "") if dato[4] != "NULL" else 0)
            facturaEncabezadoQuery.precioVentaA = round(float(
                dato[7].replace('"', "") if dato[7] != "NULL" else 0), 2)
            facturaEncabezadoQuery.precioVentaB = round(float(
                dato[8].replace('"', "") if dato[8] != "None" else 0), 2)
            facturaEncabezadoQuery.precioVentaC = round(float(
                dato[9].replace('"', "") if dato[9] != "None" else 0), 2)
            facturaEncabezadoQuery.precioVentaD = round(float(
                dato[10].replace('"', "") if dato[10] != "None" else 0), 2)
            facturaEncabezadoQuery.precioVentaE = round(float(
                dato[11].replace('"', "") if dato[11] != "None" else 0), 2)
            facturaEncabezadoQuery.updated_at = str(timezone_now)
            print('SALE IF', facturaEncabezadoQuery)
            facturaEncabezadoQuery.save()
            return 'Dato insertado correctamente'
        else:
            return f'Error con el codigo: {dato[0]}'
    except Exception as e:
        print('error', e)
        return str(e)


def insertarDato_StockProductoMegabahia(dato, resetearStock):
    print('dato', dato)
    try:
        timezone_now = timezone.localtime(timezone.now())
        if dato[0] == None and dato[4] == None or dato[0] == 'None' and dato[4] == 'None':
            return 'Dato insertado correctamente'
        facturaEncabezadoQuery = ProductosMDP.objects.filter(codigoBarras=str(dato[0]), canal='megabahia.megadescuento.com').first()
        print('encontro', facturaEncabezadoQuery)
        if facturaEncabezadoQuery and None is not facturaEncabezadoQuery:
            print('Entro al if', resetearStock)
            if resetearStock:
                facturaEncabezadoQuery.stock = int(dato[3].replace('"', "") if dato[3] != "NULL" else 0)
            else:
                facturaEncabezadoQuery.stock = facturaEncabezadoQuery.stock + int(
                    dato[3].replace('"', "") if dato[3] != "NULL" else 0)
            facturaEncabezadoQuery.precioVentaA = round(float(
                dato[5].replace('"', "") if dato[5] != "NULL" else 0), 2)
            facturaEncabezadoQuery.precioVentaB = round(float(
                dato[6].replace('"', "") if dato[6] != "NULL" else 0), 2)
            facturaEncabezadoQuery.precioVentaC = round(float(
                dato[7].replace('"', "") if dato[7] != "NULL" else 0), 2)
            facturaEncabezadoQuery.precioVentaD = round(float(
                dato[8].replace('"', "") if dato[8] != "NULL" else 0), 2)
            facturaEncabezadoQuery.precioVentaE = round(float(
                dato[9].replace('"', "") if dato[9] != "NULL" else 0), 2)
            facturaEncabezadoQuery.updated_at = str(timezone_now)
            facturaEncabezadoQuery.canal = 'megabahia.megadescuento.com'
            facturaEncabezadoQuery.save()
            return 'Dato insertado correctamente'
        else:
            return f'Error con el codigo: {dato[0]}'
    except Exception as e:
        print('error', e)
        return str(e)



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def sincronizar_fotos_productos(request):
    """
    Este metodo realiza una sincronizacion de las fotos de los productos subidos
    @rtype: No DEvuelve
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'create/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'GET':
        try:
            logModel['dataEnviada'] = str(request.data)

            productos = ProductosMDP.objects.all()

            for producto in productos:
                fotoFrente = aws_s3_instancia.get_foto_frente_url(producto.codigoBarras)
                fotoBonita = aws_s3_instancia.get_foto_bonita_url(producto.codigoBarras)
                fotoOriginal = aws_s3_instancia.get_foto_original_url(producto.codigoBarras)

                if ProductosImagenesMDP.objects.filter(imagen=fotoOriginal).first() is None:
                    if fotoOriginal:
                        print('entro if')
                        ProductosImagenesMDP.objects.create(**{
                            "producto": producto,
                            "imagen": fotoOriginal,
                            "created_at": str(timezone_now)
                        })
                if ProductosImagenesMDP.objects.filter(imagen=fotoBonita).first() is None:
                    if fotoBonita:
                        print('entro if')
                        ProductosImagenesMDP.objects.create(**{
                            "producto": producto,
                            "imagen": fotoBonita,
                            "created_at": str(timezone_now)
                        })
                if ProductosImagenesMDP.objects.filter(imagen=fotoFrente).first() is None:
                    if fotoFrente:
                        print('entro if')
                        ProductosImagenesMDP.objects.create(**{
                            "producto": producto,
                            "imagen": fotoFrente,
                            "created_at": str(timezone_now)
                        })
            # createLog(logModel, serializer.errors, logExcepcion)
            return Response({'Error'}, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)
