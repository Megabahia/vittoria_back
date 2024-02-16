from rest_framework import status
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
import requests
from datetime import datetime
from ...config.config import FAC_URL, FAC_USER, FAC_PASS, FAC_AMBIENTE
# excel
import openpyxl
# logs
from ...ADM.vittoria_logs.methods import createLog, datosTipoLog, datosProductosMDP

from .serializers import ArchivosFacturasSerializer, FacturasListarSerializer, FacturasSerializer
from .models import FacturasEncabezados, FacturasDetalles

# declaracion variables log
datosAux = datosProductosMDP()
datosTipoLogAux = datosTipoLog()
# asignacion datos modulo
logModulo = datosAux['modulo']
logApi = datosAux['api']
# asignacion tipo de datos
logTransaccion = datosTipoLogAux['transaccion']
logExcepcion = datosTipoLogAux['excepcion']


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def uploadEXCEL_subirFacturas(request):
    """
    El metodo sirve para crear
    @type request: REcibe los campos de la tabla credito archivo
    @rtype: DEvuelve el registro creado
    """
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'create/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'CREAR',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            request.data['created_at'] = str(timezone_now)
            if 'updated_at' in request.data:
                request.data.pop('updated_at')

            serializer = ArchivosFacturasSerializer(data=request.data)
            uploadEXCEL_crearProductos(request)
            if serializer.is_valid():
                serializer.save()
                createLog(logModel, serializer.data, logTransaccion)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            createLog(logModel, serializer.errors, logExcepcion)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


def uploadEXCEL_crearProductos(request):
    contValidos = 0
    contInvalidos = 0
    contTotal = 0
    errores = []
    try:
        if request.method == 'POST':
            first = True  # si tiene encabezado
            uploaded_file = request.FILES['archivo']
            # you may put validations here to check extension or file size
            wb = openpyxl.load_workbook(uploaded_file)
            # getting a particular sheet by name out of many sheets
            worksheet = wb["Pedidos"]
            lines = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            lines.append(row_data)

        for dato in lines:
            contTotal += 1
            if first:
                first = False
                continue
            else:
                if worksheet.iter_cols():
                    resultadoInsertar = insertarDato_Factura(dato)
                    if resultadoInsertar != 'Dato insertado correctamente':
                        if resultadoInsertar in 'Codigo producto':
                            contInvalidos += 1
                            errores.append(
                                {"error": "Producto no encontrado " + str(contTotal) + ": " + str(resultadoInsertar)})
                        else:
                            contInvalidos += 1
                            errores.append(
                                {"error": "Error en la línea " + str(contTotal) + ": " + str(resultadoInsertar)})
                    else:
                        contValidos += 1
                else:
                    contInvalidos += 1
                    errores.append({"error": "Error en la línea " + str(
                        contTotal) + ": la fila tiene un tamaño incorrecto (" + str(len(dato)) + ")"})

        result = {"mensaje": "La Importación se Realizo Correctamente",
                  "correctos": contValidos,
                  "incorrectos": contInvalidos,
                  "errores": errores
                  }
        return Response(result, status=status.HTTP_201_CREATED)

    except Exception as e:
        err = {"error": 'Error verifique el archivo, un error ha ocurrido: {}'.format(e)}
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


# INSERTAR DATOS EN LA BASE INDIVIDUAL
def insertarDato_Factura(dato):
    try:
        timezone_now = timezone.localtime(timezone.now())
        print('llego insertar')
        facturaEncabezadoQuery = FacturasEncabezados.objects.filter(numeroPedido=dato[0]).first()
        print('facturaEncabezado', facturaEncabezadoQuery)
        if facturaEncabezadoQuery is None:
            facturaEncabezado = {}
            facturaEncabezado['numeroPedido'] = dato[0].replace('"', "") if dato[0].replace('"', "") != "NULL" else None
            facturaEncabezado['estadoPedido'] = dato[1].replace('"', "") if dato[1] != "NULL" else None
            facturaEncabezado['fechaPedido'] = str(dato[2].replace('"', "")[:10]) if dato[2] != "NULL" else None
            facturaEncabezado['notaCliente'] = dato[3].replace('"', "") if dato[3] != "NULL" else None
            facturaEncabezado['nombresFacturacion'] = dato[4].replace('"', "") if dato[4] != "NULL" else None
            facturaEncabezado['apellidosFacturacion'] = dato[5].replace('"', "") if dato[5] != "NULL" else None
            facturaEncabezado['empresaFacturacion'] = dato[6].replace('"', "") if dato[6] != "NULL" else None
            facturaEncabezado['direccionFacturacion'] = dato[7].replace('"', "") if dato[7] != "NULL" else None
            facturaEncabezado['ciudadFacturacion'] = dato[8].replace('"', "") if dato[8] != "NULL" else None
            facturaEncabezado['provinciaFacturacion'] = dato[9].replace('"', "") if dato[9] != "NULL" else None
            facturaEncabezado['codigoPostalFacturacion'] = dato[10].replace('"', "") if dato[10] != "NULL" else None
            facturaEncabezado['paisFacturacion'] = dato[11].replace('"', "") if dato[11] != "NULL" else None
            facturaEncabezado['correoElectronicoFacturacion'] = dato[12].replace('"', "") if dato[
                                                                                                 12] != "NULL" else None
            facturaEncabezado['telefonoFacturacion'] = dato[13].replace('"', "") if dato[13] != "NULL" else None
            facturaEncabezado['nombresEnvio'] = dato[14].replace('"', "") if dato[14] != "NULL" else None
            facturaEncabezado['apellidosEnvio'] = dato[15].replace('"', "") if dato[15] != "NULL" else None
            facturaEncabezado['direccionEnvio'] = dato[16].replace('"', "") if dato[16] != "NULL" else None
            facturaEncabezado['ciudadEnvio'] = dato[17].replace('"', "") if dato[17] != "NULL" else None
            facturaEncabezado['provinciaEnvio'] = dato[18].replace('"', "") if dato[18] != "NULL" else None
            facturaEncabezado['codigoPostalEnvio'] = dato[19].replace('"', "") if dato[19] != "NULL" else None
            facturaEncabezado['paisEnvio'] = dato[20].replace('"', "") if dato[20] != "NULL" else None
            facturaEncabezado['metodoPago'] = dato[21].replace('"', "") if dato[21] != "NULL" else None
            facturaEncabezado['descuentoCarrito'] = dato[22].replace('"', "") if dato[22] != "NULL" else None
            facturaEncabezado['subtotalPedido'] = dato[23].replace('"', "") if dato[23] != "NULL" else None
            facturaEncabezado['metodoEnvio'] = dato[24].replace('"', "") if dato[24] != "NULL" else None
            facturaEncabezado['importeEnvioPedido'] = dato[25].replace('"', "") if dato[25] != "NULL" else None
            facturaEncabezado['importeReemsolsadoPedido'] = dato[26].replace('"', "") if dato[26] != "NULL" else None
            facturaEncabezado['importeTotalPedido'] = dato[27].replace('"', "") if dato[27] != "NULL" else None
            facturaEncabezado['importeTotalImpuestoPedido'] = dato[28].replace('"', "") if dato[28] != "NULL" else None
            facturaEncabezado['created_at'] = str(timezone_now)

            facturaEncabezadoQuery = FacturasEncabezados.objects.create(**facturaEncabezado)

        facturaDetalleQuery = FacturasDetalles.objects.filter(numeroPedido=dato[0].replace('"', ""),
                                                              SKU=dato[29].replace('"', "")).first()
        print('detalle', facturaDetalleQuery)
        if facturaDetalleQuery is None:
            facturaDetalle = {}
            facturaDetalle[
                'facturaEncabezado_id'] = facturaEncabezadoQuery.id if facturaEncabezadoQuery is not None else facturaEncabezadoQuery.id
            facturaDetalle['numeroPedido'] = dato[0].replace('"', "")
            facturaDetalle['SKU'] = dato[29].replace('"', "") if dato[29] != "NULL" else None
            facturaDetalle['articulo'] = dato[30].replace('"', "") if dato[30] != "NULL" else None
            facturaDetalle['nombreArticulo'] = dato[31].replace('"', "") if dato[31] != "NULL" else None
            facturaDetalle['cantidad'] = dato[32].replace('"', "") if dato[32] != "NULL" else None
            facturaDetalle['precio'] = dato[33].replace('"', "") if dato[33] != "NULL" else None
            facturaDetalle['cupon'] = dato[34].replace('"', "") if dato[34] != "NULL" else None
            facturaDetalle['importeDescuento'] = dato[35].replace('"', "") if dato[35] != "NULL" else None
            facturaDetalle['importeImpuestoDescuento'] = dato[36].replace('"', "") if dato[36] != "NULL" else None
            facturaDetalle['created_at'] = str(timezone_now)

            print('antes de guardar', facturaDetalle)
            FacturasDetalles.objects.create(**facturaDetalle)
            print('se guardar')

        return 'Dato insertado correctamente'
    except Exception as e:
        return str(e)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def facturas_list(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)
            # paginacion
            page_size = int(request.data['page_size'])
            page = int(request.data['page'])
            offset = page_size * page
            limit = offset + page_size
            # Filtros
            filters = {"state": "1"}
            if 'negocio' in request.data:
                if request.data['negocio'] != '':
                    filters['negocio'] = request.data['negocio']
            if 'cliente' in request.data:
                if request.data['cliente'] != '':
                    filters['cliente'] = request.data['cliente']
            # if 'cedula' in request.data:
            #     if request.data['cedula']!='':
            #         filters['cedula'] = str(request.data['cedula'])
            # if 'inicio' and 'fin' in request.data:
            #     # if request.data['inicio'] !='':
            #     #     filters['created_at__startswith'] = str(request.data['inicio'])
            #     if request.data['inicio'] and request.data['fin'] !='':
            #         filters['created_at__range'] = [str(request.data['inicio']),str(request.data['fin'])]

            # Serializar los datos
            query = FacturasEncabezados.objects.filter(**filters).order_by('-created_at')
            serializer = FacturasSerializer(query[offset:limit], many=True)
            new_serializer_data = {'cont': query.count(),
                                   'info': serializer.data}
            # envio de datos
            return Response(new_serializer_data, status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def factura_findOne(request, pk):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'listOne/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    try:
        try:
            query = FacturasEncabezados.objects.get(pk=pk, state=1)
        except FacturasEncabezados.DoesNotExist:
            err = {"error": "No existe"}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_404_NOT_FOUND)
        # tomar el dato
        if request.method == 'GET':
            serializer = FacturasSerializer(query)
            createLog(logModel, serializer.data, logTransaccion)
            return Response(serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        err = {"error": 'Un error ha ocurrido: {}'.format(e)}
        createLog(logModel, err, logExcepcion)
        return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def factura_facturar(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)

            resp = requests.post(f"{FAC_URL}/auth/login", json={
                "username": FAC_USER,
                "pass": FAC_PASS
            }, verify=False)
            token = resp.json()['data']['token']
            emissionPointId = resp.json()['data']['user']['accounts'][0]['id']
            enviarFacturaExternas(emissionPointId, token, request.data['facturas'])

            # envio de datos
            return Response('Ok', status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def factura_facturar_locales(request):
    timezone_now = timezone.localtime(timezone.now())
    logModel = {
        'endPoint': logApi + 'list/',
        'modulo': logModulo,
        'tipo': logExcepcion,
        'accion': 'LEER',
        'fechaInicio': str(timezone_now),
        'dataEnviada': '{}',
        'fechaFin': str(timezone_now),
        'dataRecibida': '{}'
    }
    if request.method == 'POST':
        try:
            logModel['dataEnviada'] = str(request.data)

            resp = requests.post(f"{FAC_URL}/auth/login", json={
                "username": FAC_USER,
                "pass": FAC_PASS
            }, verify=False)
            token = resp.json()['data']['token']
            emissionPointId = resp.json()['data']['user']['accounts'][0]['id']
            enviarFacturaLocales(emissionPointId, token, request.data['facturas'])

            # envio de datos
            return Response('Ok', status=status.HTTP_200_OK)
        except Exception as e:
            err = {"error": 'Un error ha ocurrido: {}'.format(e)}
            createLog(logModel, err, logExcepcion)
            return Response(err, status=status.HTTP_400_BAD_REQUEST)


def enviarFacturaExternas(emissionPointId, token, facturas):
    for item in facturas:
        detallesFacturaEnviar = []
        for itemDetalles in item['detalles']:
            precio = round(float(itemDetalles['precio']) / 1.12, 2) if 'precio' in itemDetalles else 0
            precio = precio * int(itemDetalles['cantidad'])
            ivaItem = round(float(precio) * 0.12, 2)
            detallesFacturaEnviar.append({
                "codigoPrincipal": "1",
                "descripcion": itemDetalles['nombreArticulo'],
                "cantidad": itemDetalles['cantidad'],
                "precioUnitario": itemDetalles['precio'],
                "descuento": itemDetalles['importeDescuento'] if 'importeDescuento' in itemDetalles and itemDetalles['importeDescuento'] != 'None' else 0,
                "precioTotalSinImpuesto": precio,
                "impuestos": [
                    {
                        "codigo": "2",
                        "codigoPorcentaje": "2",
                        "tarifa": "12",
                        "baseImponible": precio,
                        "valor": ivaItem
                    }
                ]
            })

        subTotal = round(float(item['subtotalPedido']) / 1.12, 2)
        ivaTotal = round(float(subTotal) * 0.12, 2)
        # Convertir a objeto datetime
        fecha_objeto = datetime.strptime(item['fechaPedido'], "%Y-%m-%d")

        # Formatear como dd/mm/yyyy
        fecha_formateada = fecha_objeto.strftime("%d/%m/%Y")
        enviarFactura = {
            "emissionPointId": emissionPointId,
            "receiptTypeId": "1",
            "parishId": "1",
            "receipt": {
                "infoTributaria": {
                    "ambiente": FAC_AMBIENTE,
                    "tipoEmision": "1"
                },
                "infoFactura": {
                    "fechaEmision":  fecha_formateada,
                    "razonSocialComprador": f"{item['nombresFacturacion']} {item['apellidosFacturacion']}",
                    "tipoIdentificacionComprador": '05',
                    "identificacionComprador": '1003150602',
                    "direccionComprador": item['direccionFacturacion'],
                    "totalSinImpuestos": subTotal,
                    "totalDescuento": item['descuentoCarrito'],
                    "totalConImpuestos": [
                        {
                            "codigo": "2",
                            "codigoPorcentaje": "2",
                            "baseImponible": subTotal,
                            "valor": ivaTotal
                        }
                    ],
                    "propina": "0",
                    "importeTotal": float(item['subtotalPedido']),
                    "moneda": "DOLAR",
                    "pagos": [
                        {
                            "formaPago": "20",
                            "total": float(item['subtotalPedido'])
                        }
                    ]
                },
                "detalles": detallesFacturaEnviar,
                "infoAdicional": [
                    {
                        "nombre": "Email",
                        "valor": item['correoElectronicoFacturacion']
                    },
                    {
                        "nombre": "Email2",
                        "valor": item['correoElectronicoFacturacion']
                    },
                    {
                        "nombre": "numeroPedido",
                        "valor": item['numeroPedido']
                    }
                ]
            }
        }

        # print('facura enviar', enviarFactura)
        resp = requests.post(f"{FAC_URL}/receipt/create", headers={"Authorization": token}, json=enviarFactura,
                             verify=False)
        print('facturas', resp.json())


def enviarFacturaLocales(emissionPointId, token, facturas):
    for item in facturas:
        detallesFacturaEnviar = []
        for itemDetalles in item['detalles']:
            precio = round(float(itemDetalles['valorUnitario']), 2)
            precio = precio * int(itemDetalles['cantidad'])
            ivaItem = round(float(precio) * 0.12, 2)
            detallesFacturaEnviar.append({
                "codigoPrincipal": "1",
                "descripcion": itemDetalles['articulo'],
                "cantidad": itemDetalles['cantidad'],
                "precioUnitario": itemDetalles['valorUnitario'],
                "descuento": itemDetalles['descuento'],
                "precioTotalSinImpuesto": precio,
                "impuestos": [
                    {
                        "codigo": "2",
                        "codigoPorcentaje": "2",
                        "tarifa": "12",
                        "baseImponible": precio,
                        "valor": ivaItem
                    }
                ]
            })

        subTotal = round(float(item['subTotal']), 2)
        ivaTotal = round(float(item['iva']), 2)
        # Convertir a objeto datetime
        fecha_objeto = datetime.strptime(item['fecha'], "%Y-%m-%d")

        # Formatear como dd/mm/yyyy
        fecha_formateada = fecha_objeto.strftime("%d/%m/%Y")
        enviarFactura = {
            "emissionPointId": emissionPointId,
            "receiptTypeId": "1",
            "parishId": "1",
            "receipt": {
                "infoTributaria": {
                    "ambiente": FAC_AMBIENTE,
                    "tipoEmision": "1"
                },
                "infoFactura": {
                    "fechaEmision": fecha_formateada,
                    "razonSocialComprador": f"{item['cliente']['nombres']} {item['cliente']['apellidos']}",
                    "tipoIdentificacionComprador": '05',
                    "identificacionComprador": item['cliente']['cedula'],
                    "direccionComprador": item['direccion'],
                    "totalSinImpuestos": subTotal,
                    "totalDescuento": item['descuento'],
                    "totalConImpuestos": [
                        {
                            "codigo": "2",
                            "codigoPorcentaje": "2",
                            "baseImponible": subTotal,
                            "valor": ivaTotal
                        }
                    ],
                    "propina": "0",
                    "importeTotal": item['total'],
                    "moneda": "DOLAR",
                    "pagos": [
                        {
                            "formaPago": "20",
                            "total": item['total']
                        }
                    ]
                },
                "detalles": detallesFacturaEnviar,
                "infoAdicional": [
                    {
                        "nombre": "Email",
                        "valor": item['correo']
                    },
                    {
                        "nombre": "Email2",
                        "valor": item['correo']
                    }
                ]
            }
        }

        print('enviarFactura', enviarFactura)

        resp = requests.post(f"{FAC_URL}/receipt/create", headers={"Authorization": token}, json=enviarFactura,
                             verify=False)
        print('facturas', resp.json())
